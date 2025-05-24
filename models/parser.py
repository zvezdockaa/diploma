import requests
import openpyxl

# парсер для получения данных из всемирного банка и файла с военными расходами
class Parser:
    def __init__(self, military_spending_file="военные расходы1.xlsx", sheet_name="Current US$"):
        self.military_spending_file = military_spending_file
        self.sheet_name = sheet_name

    # функция получает значение индикатора по коду страны и году
    def fetch_data_from_world_bank(self, country_code, indicator, year=2023):
        url = f"http://api.worldbank.org/v2/country/{country_code}/indicator/{indicator}?format=json&date={year}"
        try:
            response = requests.get(url)
            response.raise_for_status()
            data = response.json()

            if len(data) < 2 or not data[1]:
                return None

            # возвращаем первое доступное значение
            for entry in data[1]:
                value = entry.get("value")
                if value is not None:
                    return value
            return None

        except requests.exceptions.RequestException as e:
            print(f"[Parser] Ошибка при получении данных с World Bank API: {e}")
            return None

    # функция получает военные расходы из файла по названию страны и году
    def fetch_military_spending(self, country_name, year):
        try:
            workbook = openpyxl.load_workbook(self.military_spending_file, data_only=True)
            sheet = workbook[self.sheet_name]

            # заголовки — в шестой строке
            header_row = next(sheet.iter_rows(min_row=6, max_row=6, values_only=True))
            year_col_index = None

            # определяем нужный столбец по году
            for idx, cell in enumerate(header_row):
                if str(cell).strip() == str(year):
                    year_col_index = idx
                    break

            if year_col_index is None:
                return 0, f"Год {year} не найден в таблице."

            # ищем строку с нужной страной, начиная с 7-й строки
            for row in sheet.iter_rows(min_row=7, values_only=True):
                if not row or not row[0]:
                    continue

                country_cell = str(row[0]).strip()
                notes_cell = str(row[1]).strip() if row[1] else ""

                # пропускаем подзаголовки и пустые строки
                if notes_cell == "" and all(v in ["", None, "..."] for v in row[2:]):
                    continue

                if country_cell.lower() == country_name.strip().lower():
                    value = row[year_col_index]
                    if value is None or str(value).strip() in ["...", "nan", ""]:
                        return 0, f"Нет данных о военных расходах для '{country_name}' за {year}"
                    try:
                        return round(float(str(value).replace(",", "")), 5), None
                    except ValueError:
                        return 0, f"Невозможно интерпретировать значение для '{country_name}' за {year}'"

            return 0, f"Страна '{country_name}' не найдена в таблице"

        except Exception as e:
            raise ValueError(f"[Parser] Ошибка при открытии файла: {e}")
